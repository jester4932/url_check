import re

from openpyxl import load_workbook


class XLSXRead:
    def __init__(self, **kwargs):
        self.file = kwargs.get("file", "")
        self.regex = kwargs.get("regex", "")
        self.url = []

    def main(self):
        self.xlsx_read()
        return self.url

    def xlsx_read(self):  # finds plain text urls and hyperlinks
        excel_data = load_workbook(self.file, data_only=True)
        for sheet in excel_data.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        url = re.findall(self.regex, str(cell.value))
                        if url:
                            self.url.append(cell.value)
                    if cell.hyperlink:
                        self.url.append(cell.hyperlink.target)
